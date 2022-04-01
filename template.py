# import xlsxwriter module
import xlsxwriter 
# Reading an excel file using Python
import xlrd
import openpyxl 
# imort pandas module
import pandas as pd 
# import os module
import os

# Folder Path
path = "data"
  
# Change the directory
os.chdir(path)

workbook = xlsxwriter.Workbook('Example.xlsx')





# Read  File
def read_xlsx_file(file):

    sheet_name = file.split('F')[0]

    wb = workbook
    worksheet = workbook.add_worksheet(sheet_name)

    write_file_name_header(wb,worksheet,file)
    file_path = f"../{path}/{file}"
    wb_obj = openpyxl.load_workbook(file_path,data_only=True) 
    sheet_obj = wb_obj.active 

    # Getting the value of maximum rows
    # and column
    row = sheet_obj.max_row
    column = sheet_obj.max_column


    #Write the Block
    write_Block(worksheet,sheet_obj,row,column)

    #Write the Course
    write_Course(worksheet,sheet_obj,row,column)

    #Write the Title
    write_Title(worksheet,sheet_obj,row,column)

    #Write Instructor
    write_Instructor(worksheet,sheet_obj,row,column)

    #Write Day and Time
    write_Day_Time(worksheet,sheet_obj,row,column)

    #Write Size
    write_Size(worksheet,sheet_obj,row,column)

    #Write Location
    write_Location(worksheet,sheet_obj,row,column)

    #Write Comments
    write_Comments(worksheet,sheet_obj,row,column)




def write_file_name_header(workbook,worksheet,file):
    file_name = file.split('.')[0]

    worksheet.write(0, 0, file_name, workbook.add_format({'bold': True, 'color': '#E26B0A', 'size': 42}))
    # Start from the first cell.
    # Rows and columns are zero indexed.
    row = 1
    column = 0
    header = ["Block", "Dep", "Course", "Section","Title","Instructor","Day","Start","End","Size","Bldg",
                    "Loc", "Rm", "Comments"]

    header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'fg_color': 'black', 'border': 1,'font_color': 'white','font_size': 16})
    worksheet.set_column(0, len(header), 15)
    worksheet.set_column(4, 4, 35)
    worksheet.set_column(13, 13, 30)


    # iterating through content list
    for item in header:
    
        # write operation perform
        worksheet.write(row, column, item,header_format)
    
        # incrementing the value of column by one
        # with each iterations.
        column += 1    
            

def write_Block(worksheet,sheet_obj,row,column):
    cell_obj = sheet_obj.cell(row = 3, column = 1) 
    if str(cell_obj.value) == "Block":
        for i in range(4,get_number_of_records(sheet_obj,row,column)): 
            cell_obj = sheet_obj.cell(row = i, column = 1)
            worksheet.write(i-2, 0, str(fix_Block(cell_obj.value))) 



def write_Course(worksheet,sheet_obj,row,column):
    for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Course":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 1, str(cell_obj.value))

            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c+1) 
                worksheet.write(i-2, 2, str(cell_obj.value))

            for i in range(4,get_number_of_records(sheet_obj,row,column)):
                cell_obj = sheet_obj.cell(row = i, column = c+2) 
                worksheet.write(i-2, 3, str(cell_obj.value)) 

   
def write_Title(worksheet,sheet_obj,row,column):
    for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Title":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 4, str(cell_obj.value))

def write_Instructor(worksheet,sheet_obj,row,column):
    for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Instructor":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 5, str(cell_obj.value))

def write_Day_Time(worksheet,sheet_obj,row,column):
    for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Day":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 6, str(cell_obj.value))

            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c+1)
                worksheet.write(i-2, 7, str(fix_time_format(cell_obj.value)))

            for i in range(4,get_number_of_records(sheet_obj,row,column)):
                cell_obj = sheet_obj.cell(row = i, column = c+2) 
                worksheet.write(i-2, 8, str(fix_time_format(cell_obj.value))) 

def write_Size(worksheet,sheet_obj,row,column):
    for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Max Size":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 9, str(cell_obj.value))

def write_Location(worksheet,sheet_obj,row,column):
     for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Bldg":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 10, str(fix_location(cell_obj.value)))

            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c+1)
                worksheet.write(i-2, 11, str(fix_location(cell_obj.value)))

            for i in range(4,get_number_of_records(sheet_obj,row,column)):
                cell_obj = sheet_obj.cell(row = i, column = c+2) 
                worksheet.write(i-2, 12, str(fix_location(cell_obj.value))) 

def write_Comments(worksheet,sheet_obj,row,column):
    for c in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = c) 
        if str(cell_obj.value) == "Comment I":
            for i in range(4,get_number_of_records(sheet_obj,row,column)): 
                cell_obj = sheet_obj.cell(row = i, column = c) 
                worksheet.write(i-2, 13, str(fix_comments(cell_obj.value)))

##############UTILITY##########################

def fix_time_format(value):
    if  str(value).endswith("AM"):
        value = str(value)[:-2] + " AM"
                
    if  str(value).endswith("PM"):
        value = str(value)[:-2] + " PM"

    if  str(value).endswith("am"):
        value = str(value)[:-2] + " AM"

    if  str(value).endswith("pm"):
         value = str(value)[:-2] + " PM"

    if  str(value).endswith("N/A"):
        value = str(value)[:-4] + "Online"

    if  str(value).endswith(":00"):
        if int(str(value)[:2]) <=12:
            value = str(value)[:-3] + " AM"

    if  str(value).endswith(":00"):
        if int(str(value)[:2]) >=13:
            hours = int(str(value)[:2]) -12
            value = str(hours) + (str(value)[2:])
            value = str(value)[:-3] + " PM"

    return value

def fix_Block(value):
    if  str(value).endswith("NO#"):
        value = str(value)[:-3] + "None" 
    if  str(value).endswith(" "):
        value = str(value)[:-3] + "None" 
    if  str(value).endswith("z"):
        value = str(value)[:-1] + "Z" 
    return value

def fix_location(value):
    if  str(value).endswith("N/A"):
        value = str(value)[:-3] + "Online" 
    if  str(value).endswith("None"):
        value = str(value)[:-4] + "Online" 
    return value

def fix_comments(value):
    if  str(value).endswith(" "):
        value = str(value)[:-3] + "None" 
    return value

def get_number_of_records(sheet_obj,row,column):
    for i in range(4, row + 1): 
        cell_obj = sheet_obj.cell(row = i, column = get_title_row_number(sheet_obj,row,column)) 
        if cell_obj.value == "#N/A":
            return(i)
        if cell_obj.value == "None":
            return(i)


def get_title_row_number(sheet_obj,row,column):
     for i in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 3, column = i) 
        if str(cell_obj.value) == "Title":
            return i

def remove(sheet, row):
    # iterate the row object
    for cell in row:
          # check the value of each cell in
        # the row, if any of the value is not
        # None return without removing the row
        if cell.value == "SAMPLE":
            sheet.delete_rows(row[0].row, 1)
            break


    # get the row number from the first cell
    # and remove the row

if __name__ == '__main__':

    # iterate through all file
    for file in os.listdir():
        # Check whether file is in text format or not
        if file.endswith(".xlsx"):
            # call read xlms file function
            if file.startswith("~"):
                continue
            read_xlsx_file(file)
            
        else:
            print(file + "not xlsx file") 

    os.chdir('../')  
    workbook.close()

    
    wb = openpyxl.load_workbook("Example.xlsx")
    ws = wb.active

    file = 'Master.xlsx'
    wb2 = xlsxwriter.Workbook(file)
    ws2 = wb2.add_worksheet()
    write_file_name_header(wb2,ws2,file)

  
    

    var = 0
    # first_shhet = mr

    for ws in wb.worksheets:
        mr = ws.max_row
        mc = ws.max_column


        
         # iterate the sheet object and remove sample
        for row in ws:
            remove(ws,row)
        
        # copying the cell values from source 
        # excel file to destination excel file
        for i in range (3, mr + 1):
            for j in range (1, mc + 1):
                # reading cell value from source excel file
                c = ws.cell(row = i, column = j)
                var2 = var + i
                # writing the read value to destination excel file
                ws2.write(var2-1,j-1,c.value)
        var += mr
        var2 = 0
    wb.save('Example.xlsx')
    wb.close()
    wb2.close()
               


      

